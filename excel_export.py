import os
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


KOLONE = [
    "redni_broj", "naziv", "kategorija", "adresa", "telefon", "email",
    "web_stranica", "radno_vrijeme", "status", "ocjena",
    "broj_recenzija", "plus_code", "opis", "google_maps_url",
    "datum_scrape", "pozvano", "zainteresovan",
]

ZAGLAVLJE = {
    "redni_broj":     "Br.",
    "naziv":          "Naziv",
    "kategorija":     "Kategorija",
    "adresa":         "Adresa",
    "telefon":        "Telefon",
    "email":          "Email",
    "web_stranica":   "Web stranica",
    "radno_vrijeme":  "Radno vrijeme",
    "status":         "Status",
    "ocjena":         "Ocjena",
    "broj_recenzija": "Recenzije",
    "plus_code":      "Plus Code",
    "opis":           "Opis",
    "google_maps_url": "Google Maps URL",
    "datum_scrape":   "Datum scrape",
    "pozvano":        "Pozvano (DA/NE)",
    "zainteresovan":  "Zainteresovan (DA/NE)",
}

SIRINE = {
    "redni_broj": 5, "naziv": 35, "kategorija": 20, "adresa": 35,
    "telefon": 16, "email": 25, "web_stranica": 35, "radno_vrijeme": 40, "status": 14,
    "ocjena": 8, "broj_recenzija": 10, "plus_code": 14, "opis": 40,
    "google_maps_url": 45, "datum_scrape": 16,
    "pozvano": 16, "zainteresovan": 20,
}


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "", name)
    return name[:31]


def safe_filename(name: str) -> str:
    name = re.sub(r"[^\w\s\-]", "", name, flags=re.UNICODE)
    name = name.strip().replace(" ", "_")
    return name or "rezultati"


def format_sheet(ws):
    header_fill  = PatternFill("solid", fgColor="2F5597") # Lepša plava
    header_font  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    
    body_font    = Font(name="Segoe UI", size=10)
    link_font    = Font(name="Segoe UI", size=10, color="0563C1", underline="single") # Plava boja za linkove
    
    alt_fill     = PatternFill("solid", fgColor="F2F2F2") # Svetlo siva za svaki drugi red (Zebra)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    wrap_align   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    
    thin         = Side(style="thin", color="D9D9D9")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Stil zaglavlja
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = border

    # Stil tijela i Zebra boje
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_even = (row_idx % 2 == 0)
        
        for cell in row:
            idx = cell.column - 1
            key = KOLONE[idx] if idx < len(KOLONE) else ""
            
            # Ako je email, web stranica ili mapa, stavi plavi font
            if key in ("email", "web_stranica", "google_maps_url") and cell.value:
                cell.font = link_font
                if str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
            else:
                cell.font = body_font
                
            # Zebra bojenje (svaki drugi red)
            if is_even:
                cell.fill = alt_fill
                
            cell.border = border
            
            # Poravnanje
            if key in ("redni_broj", "ocjena", "broj_recenzija", "status", "pozvano", "zainteresovan"):
                cell.alignment = center_align
            elif key in ("radno_vrijeme", "opis", "google_maps_url"):
                cell.alignment = wrap_align
            else:
                cell.alignment = left_align

    # Sirine kolona
    for i, key in enumerate(KOLONE, start=1):
        ws.column_dimensions[get_column_letter(i)].width = SIRINE.get(key, 15)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 25
    
    # Dodaj AutoFilter na zaglavlje
    max_col_letter = get_column_letter(len(KOLONE))
    ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"


def save_to_excel(results_map: dict, output_dir: str, skip_duplicates: bool = True):
    """
    results_map je oblika:
    { (lokacija, sekcija, kategorija): [ lista_rezultata ], ... }
    
    Kreira folder po lokaciji, excel fajl po sekciji, sheet po kategoriji.
    """
    if not EXCEL_AVAILABLE:
        print("[UPOZORENJE] openpyxl nije instaliran. Ne mogu snimiti Excel.")
        return

    os.makedirs(output_dir, exist_ok=True)
    
    # Reorganizacija mape za lakse snimanje:
    # tree[lokacija][sekcija][kategorija] = results
    tree = {}
    for (lokacija, sekcija, kategorija), results in results_map.items():
        if not results:
            continue
            
        if lokacija not in tree: tree[lokacija] = {}
        if sekcija not in tree[lokacija]: tree[lokacija][sekcija] = {}
        tree[lokacija][sekcija][kategorija] = results
        
    for lokacija, sekcije in tree.items():
        parts = [p.strip() for p in lokacija.split(",")]
        if len(parts) == 2:
            grad, drzava = parts[0], parts[1]
        else:
            grad, drzava = "", parts[0]
            
        drzava_folder = os.path.join(output_dir, safe_filename(drzava))
        os.makedirs(drzava_folder, exist_ok=True)
        
        for sekcija, kategorije in sekcije.items():
            if grad:
                filename = f"{safe_filename(sekcija)}_{safe_filename(grad)}.xlsx"
            else:
                filename = f"{safe_filename(sekcija)}.xlsx"
                
            filepath = os.path.join(drzava_folder, filename)
            
            if os.path.exists(filepath):
                wb = openpyxl.load_workbook(filepath)
            else:
                wb = openpyxl.Workbook()
                default = wb.active
                if default and default.title in ("Sheet", "Sheet1"):
                    del wb[default.title]

            for kategorija, results in kategorije.items():
                sname = safe_sheet_name(kategorija)
                existing_urls = set()
                
                if sname in wb.sheetnames:
                    if not skip_duplicates:
                        del wb[sname]
                        ws = wb.create_sheet(title=sname)
                        ws.append([ZAGLAVLJE.get(k, k) for k in KOLONE])
                    else:
                        ws = wb[sname]
                        url_col_idx = KOLONE.index("google_maps_url") + 1
                        for row_idx in range(2, ws.max_row + 1):
                            val = ws.cell(row=row_idx, column=url_col_idx).value
                            if val:
                                existing_urls.add(val)
                else:
                    ws = wb.create_sheet(title=sname)
                    ws.append([ZAGLAVLJE.get(k, k) for k in KOLONE])

                added_count = 0
                for row in results:
                    url = row.get("google_maps_url", "")
                    if skip_duplicates and url in existing_urls:
                        continue
                    
                    if skip_duplicates:
                        row["redni_broj"] = ws.max_row # max_row je trenutno zadnji red
                        
                    ws.append([row.get(k, "") for k in KOLONE])
                    existing_urls.add(url)
                    added_count += 1
                
                if added_count > 0:
                    format_sheet(ws)

            wb.save(filepath)
            print(f"  [+] Sačuvano: {filepath}")
